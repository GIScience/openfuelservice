import datetime
import logging
from pathlib import Path
from typing import Any, Dict, List, Tuple, Union

import openpyxl as xl
import xlrd
from openpyxl.cell.read_only import EmptyCell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session
from tqdm import tqdm

from app.db.importer.base_reader import BaseReader
from app.db.importer.mappings import EurostatSheetMapping
from app.misc import file_management
from app.models import CountryData, EurostatCountryPrice, EurostatGeneralPrice

logger = logging.getLogger(__name__)


class EurostatFuelReader(BaseReader):
    def __init__(self, file_to_read: Union[str, Path], db: Session):
        super().__init__(file_to_read)
        self._objects_list_ordered: Dict = {}
        self._db = db
        self.name = "EurostatFuelReader"

    @staticmethod
    def _convert_string_prices_to_float(string_value: Any) -> Union[float, None]:
        float_string = str(string_value)
        if not len(float_string):
            return None
        if not any(char.isdigit() for char in float_string):
            return None
        try:
            return float(float_string.replace(",", ""))
        except Exception:
            return None

    def _convert_datetuple_to_date(
        self, date_tuple: float, datemode: int
    ) -> Union[datetime.date, None]:
        if date_tuple is None or type(date_tuple) is not float:
            return None
        try:
            date_as_datetime = datetime.datetime(
                *xlrd.xldate_as_tuple(date_tuple, datemode)
            ).date()
            return date_as_datetime
        except TypeError as err:
            logger.error(
                f"TypeError converting value: {date_tuple} with datemode: {datemode} to datetime."
            )
            raise err
        except Exception as err:
            logger.error(
                f"Unknown error converting value: {date_tuple} with datemode: {datemode} to datetime."
            )
            raise err

    def _convert_value_to_country_code(
        self, country_code: str
    ) -> Union[CountryData, None]:
        if (
            country_code is None
            or type(country_code) is not str
            or not len(country_code)
        ):
            return None
        elif len(country_code) > 3:
            # No country value greater than 3!
            return None
        country_data_match: CountryData = self._db.query(CountryData).filter(
            CountryData.country_alpha_2.in_([country_code])
        ).first()
        if country_data_match is None:
            country_data_match = (
                self._db.query(CountryData)
                .filter(CountryData.country_alpha_3.in_([country_code]))
                .first()
            )
            if country_data_match is None:
                return None
            else:
                return country_data_match
        else:
            return country_data_match

    def _determine_rows_from_sheet(
        self,
        sheet: Worksheet,
        ignore_country_column: bool = False,
        ignore_lpg_column: bool = False,
    ) -> Tuple[int, int, int, int, int, int]:
        (
            code_column,
            date_column,
            taux_column,
            eurostat_super_column,
            eurostat_diesel_column,
            lpg_column,
        ) = (-1, -1, -1, -1, -1, -1)
        row: tuple
        for row in sheet.rows:
            # Ignore empty rows
            if all(
                isinstance(cell, EmptyCell)
                or cell is None
                or cell.value is None
                or cell.data_type != "s"
                for cell in row
            ):
                continue
            for column_index in range(sheet.max_column):
                cell: Union[EmptyCell] = row[column_index]
                if (
                    isinstance(cell, EmptyCell)
                    or cell is None
                    or cell.value is None
                    or cell.data_type != "s"
                ):
                    continue
                if not ignore_country_column and code_column == -1:
                    country_data: Union[
                        CountryData, None
                    ] = self._convert_value_to_country_code(cell.value)
                    code_column = (
                        column_index
                        if issubclass(country_data.__class__, CountryData)
                        else -1
                    )
                    continue
                harmonized_value = (
                    cell.value.lower()
                    .strip()
                    .replace(" ", "")
                    .replace("\r", "")
                    .replace("\n", "")
                )
                if date_column == -1 and harmonized_value == "date":
                    date_column = column_index
                    continue
                if taux_column == -1 and harmonized_value == "exchangerateto€":
                    taux_column = column_index
                    continue
                if (
                    eurostat_super_column == -1
                    and harmonized_value == "euro-super95(i)"
                ):
                    eurostat_super_column = column_index
                    continue
                if (
                    eurostat_diesel_column == -1
                    and harmonized_value
                    == "gasoilautomobileautomotivegasoildieselkraftstoff(i)"
                ):
                    eurostat_diesel_column = column_index
                if (
                    not ignore_lpg_column
                    and lpg_column == -1
                    and harmonized_value == "gplpourmoteurlpgmotorfuel"
                ):
                    lpg_column = column_index
            if all(
                i >= 0
                for i in [
                    code_column if not ignore_country_column else 99,
                    date_column,
                    eurostat_super_column,
                    eurostat_diesel_column,
                    lpg_column if not ignore_lpg_column else 99,
                ]
            ):
                break
        return (
            code_column,
            date_column,
            taux_column,
            eurostat_super_column,
            eurostat_diesel_column,
            lpg_column,
        )

    def _process_per_ctr(self, sheet: Worksheet) -> None:
        # Set country and lpg dependency based on sheet.
        ignore_country_column: bool = False
        ignore_lpg_column: bool = False
        if any(
            sheet.title == mapping.value
            for mapping in [
                EurostatSheetMapping.PRICES_WITH_TAXES_EU,
                EurostatSheetMapping.PRICES_WO_TAXES_EU,
            ]
        ):
            ignore_country_column = True
            ignore_lpg_column = False
        elif any(
            sheet.title == mapping.value
            for mapping in [
                EurostatSheetMapping.PRICES_WITH_TAXES_UK,
                EurostatSheetMapping.PRICES_WO_TAXES_UK,
            ]
        ):
            ignore_country_column = False
            ignore_lpg_column = True

        # Determine Rows
        (
            code_column,
            date_column,
            taux_column,
            eurostat_super_column,
            eurostat_diesel_column,
            lpg_column,
        ) = self._determine_rows_from_sheet(
            sheet, ignore_country_column, ignore_lpg_column
        )

        if any(
            i < 0
            for i in [
                code_column if not ignore_country_column else 99,
                date_column,
                eurostat_super_column,
                eurostat_diesel_column,
                lpg_column if not ignore_lpg_column else 99,
            ]
        ):  # noqa
            logger.warning(f"Couldn't determine all columns for sheet {sheet.title}")
            return
        current_country: CountryData = CountryData()
        row: tuple
        for row in tqdm(
            sheet.rows,
            total=sheet.max_row,
            unit=" rows",
            desc=f" Processing fuel prices from source: {sheet.title}",
        ):
            if all(isinstance(cell, EmptyCell) or cell.value is None for cell in row):
                continue

            country_code_object: Union[
                CountryData, None
            ] = self._convert_value_to_country_code(row[code_column].value)
            date_value: datetime.datetime = row[date_column].value
            if ignore_country_column:
                pass
            elif (
                country_code_object is None and current_country.country_alpha_2 is None
            ):
                continue
            elif (
                country_code_object is not None
                and country_code_object.country_alpha_2
                is not current_country.country_alpha_2
            ):
                current_country = country_code_object

            if date_value is None or not issubclass(
                date_value.__class__, datetime.datetime
            ):
                continue
            if ignore_country_column and "EU_GENERAL" not in self._objects_list_ordered:
                self._objects_list_ordered["EU_GENERAL"] = {}
            elif (
                not ignore_country_column
                and current_country.country_alpha_2 not in self._objects_list_ordered
            ):
                self._objects_list_ordered[current_country.country_alpha_2] = {}

            taux: Union[float, None] = self._convert_string_prices_to_float(
                row[taux_column].value
            ) if taux_column >= 0 else None
            if taux is None:
                taux = float(1)

            if (
                ignore_country_column
                and date_value in self._objects_list_ordered["EU_GENERAL"]
            ):
                eurostat_price_object = self._objects_list_ordered["EU_GENERAL"][
                    date_value
                ]
            elif (
                not ignore_country_column
                and date_value
                in self._objects_list_ordered[current_country.country_alpha_2]
            ):
                eurostat_price_object = self._objects_list_ordered[
                    current_country.country_alpha_2
                ][date_value]
            elif ignore_country_column:
                eurostat_price_object = EurostatGeneralPrice()
            else:
                eurostat_price_object = EurostatCountryPrice()
                eurostat_price_object.country_alpha_2 = current_country.country_alpha_2
                if taux is not None and type(taux) is int or type(taux) is float:
                    eurostat_price_object.taux = taux

            eurostat_price_object.date = date_value
            eurostat_price_object.euro_quantity = 1000
            eurostat_price_object.diesel_quantity = 1000
            eurostat_price_object.euro_unit = "liter"
            eurostat_price_object.diesel_unit = "liter"
            eurostat_price_object.price_in_euro = 1

            eurostat_super: Union[float, None] = self._convert_string_prices_to_float(
                row[eurostat_super_column].value
            )
            eurostat_diesel: Union[float, None] = self._convert_string_prices_to_float(
                row[eurostat_diesel_column].value
            )

            if any(
                sheet.title == mapping.value
                for mapping in [
                    EurostatSheetMapping.PRICES_WITH_TAXES_PER_CTR,
                    EurostatSheetMapping.PRICES_WITH_TAXES_UK,
                    EurostatSheetMapping.PRICES_WITH_TAXES_EU,
                ]
            ):
                if eurostat_super is not None and type(eurostat_super) is float:
                    eurostat_price_object.euro_ttc = (
                        eurostat_super / eurostat_price_object.euro_quantity * taux
                    )
                if eurostat_diesel is not None and type(eurostat_diesel) is float:
                    eurostat_price_object.diesel_ttc = (
                        eurostat_diesel / eurostat_price_object.euro_quantity * taux
                    )

            elif any(
                sheet.title == mapping.value
                for mapping in [
                    EurostatSheetMapping.PRICES_WO_TAXES_PER_CTR,
                    EurostatSheetMapping.PRICES_WO_TAXES_UK,
                    EurostatSheetMapping.PRICES_WO_TAXES_EU,
                ]
            ):
                if eurostat_super is not None and type(eurostat_super) is float:
                    eurostat_price_object.euro_ht = (
                        eurostat_super / eurostat_price_object.euro_quantity * taux
                    )
                if eurostat_diesel is not None and type(eurostat_diesel) is float:
                    eurostat_price_object.diesel_ht = (
                        eurostat_diesel / eurostat_price_object.euro_quantity * taux
                    )
            if (
                isinstance(eurostat_price_object, EurostatCountryPrice)
                and eurostat_price_object.is_valid
            ):
                self._objects_list_ordered[current_country.country_alpha_2][
                    date_value
                ] = eurostat_price_object
            elif (
                isinstance(eurostat_price_object, EurostatGeneralPrice)
                and eurostat_price_object.is_valid
            ):
                self._objects_list_ordered["EU_GENERAL"][
                    date_value
                ] = eurostat_price_object

    def _process_data(self, data_file: Path) -> None:
        files: list = file_management.unzip_download(
            zip_file_path=data_file, destination_folder=self._tempfolder
        )
        all_countries_from_db: List = self._db.query(CountryData).all()
        if not len(all_countries_from_db):
            logger.warning(
                "Stopped importing fuel prices. No country data found to connect the prices to. Import "
                "countries first."
            )
            return
        sheets = [
            EurostatSheetMapping.PRICES_WITH_TAXES_PER_CTR,
            EurostatSheetMapping.PRICES_WO_TAXES_PER_CTR,
            EurostatSheetMapping.PRICES_WITH_TAXES_EU,
            EurostatSheetMapping.PRICES_WO_TAXES_EU,
            EurostatSheetMapping.PRICES_WITH_TAXES_UK,
            EurostatSheetMapping.PRICES_WO_TAXES_UK,
        ]
        file: Path
        for file in files:
            if file.name.rsplit(".", 1)[-1] == "xls":
                continue
            elif file.name.rsplit(".", 1)[-1] == "xlsx":
                wb = xl.load_workbook(
                    file, read_only=True, keep_vba=False, data_only=True
                )
                sheet_mapping: EurostatSheetMapping
                for sheet_mapping in sheets:
                    self._process_per_ctr(wb.get_sheet_by_name(sheet_mapping.value))
        database_objects: Dict
        for key, database_objects in self._objects_list_ordered.items():
            while database_objects:
                self.objects_list.append(database_objects.popitem()[1])
